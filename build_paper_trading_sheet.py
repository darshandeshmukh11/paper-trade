#!/usr/bin/env python3
"""Generate paper_trading_india.xlsx for NSE/BSE equity paper trading."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = Path(__file__).resolve().parent / "paper_trading_india.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1E3A5F")
MONEY_FMT = '#,##0.00'
PCT_FMT = "0.00%"
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row: int, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def set_col_widths(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def build_settings(ws) -> None:
    ws.title = "Settings"
    ws["A1"] = "Indian stock market — paper trading settings"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    rows = [
        ("", "Parameter", "Value", "Notes"),
        ("", "Starting capital (₹)", 500_000, "Virtual cash for paper trading"),
        ("", "Brokerage per order (₹)", 20, "Typical discount broker flat fee"),
        ("", "GST on brokerage (%)", 0.18, "18% on brokerage"),
        ("", "STT — delivery sell (%)", 0.001, "0.1% on sell turnover"),
        ("", "STT — intraday sell (%)", 0.00025, "0.025% on sell turnover"),
        ("", "Exchange txn charge (%)", 0.0000345, "NSE equity ~0.00345% each side"),
        ("", "SEBI fee (%)", 0.000001, "₹10 per crore (approx)"),
        ("", "Stamp duty — buy (%)", 0.00015, "Varies by state; default ~0.015%"),
        ("", "DP charges — delivery sell (₹)", 15.93, "Per sell delivery scrip (approx)"),
        ("", "Currency", "INR", ""),
        ("", "Default exchange", "NSE", "NSE or BSE"),
    ]
    for i, row in enumerate(rows, start=3):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    style_header_row(ws, 3, 4)
    set_col_widths(ws, {1: 4, 2: 32, 3: 18, 4: 42})

    # Named ranges for formulas (row refs fixed)
    ws["C4"].number_format = MONEY_FMT
    ws["C5"].number_format = MONEY_FMT
    for r in (7, 8, 9, 10, 11, 12):
        ws.cell(row=r, column=3).number_format = PCT_FMT
    ws["C13"].number_format = MONEY_FMT

    ws["A16"] = "Charge formulas use Settings!$C$4:$C$13. Edit values above to match your broker."
    ws["A16"].font = Font(italic=True, color="64748B")


def build_instructions(ws) -> None:
    ws.title = "Instructions"
    lines = [
        "Paper trading workbook — Indian equities (NSE / BSE)",
        "",
        "1. Settings — Set starting capital and charge assumptions.",
        "2. Trade Log — Log every BUY and SELL (one row per fill). Use the same Position ID for entry and exit.",
        "3. Open Positions — Auto view of symbols still held (net qty > 0).",
        "4. Closed Trades — Enter exit row with Position ID; realized P&L is calculated.",
        "5. Summary — Cash, exposure, and performance metrics.",
        "",
        "Tips:",
        "• Symbol: use NSE tickers (e.g. RELIANCE, TCS, JINDALSTEL).",
        "• Segment: Equity Delivery or Equity Intraday (affects STT).",
        "• For a swing trade: BUY row Position ID = T001, SELL row same ID = T001.",
        "• LTP (last traded price) on Open Positions is manual — update for mark-to-market.",
        "",
        "Regenerate this file: python build_paper_trading_sheet.py",
        f"File location: {OUTPUT}",
    ]
    for i, line in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=line)
    ws["A1"].font = TITLE_FONT
    ws.column_dimensions["A"].width = 88


def build_trade_log(ws) -> None:
    ws.title = "Trade Log"
    headers = [
        "Trade #",
        "Date",
        "Symbol",
        "Exchange",
        "Segment",
        "Side",
        "Qty",
        "Price (₹)",
        "Position ID",
        "Notes",
        "Gross (₹)",
        "Brokerage (₹)",
        "STT (₹)",
        "Exch+SEBI (₹)",
        "Stamp (₹)",
        "GST (₹)",
        "DP (₹)",
        "Total charges (₹)",
        "Net cash flow (₹)",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    # Sample rows
    samples = [
        [1, "2026-05-27", "JINDALSTEL", "NSE", "Equity Delivery", "BUY", 50, 1205, "T001", "Support rejection entry"],
        [2, "2026-05-29", "JINDALSTEL", "NSE", "Equity Delivery", "SELL", 50, 1222, "T001", "Resistance exit"],
    ]
    start_row = 2
    max_row = 502  # 500 trade rows
    for i, sample in enumerate(samples):
        r = start_row + i
        for j, val in enumerate(sample, start=1):
            ws.cell(row=r, column=j, value=val)

    for r in range(start_row, max_row + 1):
        # Trade # auto
        ws.cell(row=r, column=1, value=f"=IF(G{r}=\"\",\"\",ROW()-1)")
        # Gross
        ws.cell(row=r, column=11, value=f'=IF(G{r}="","",G{r}*H{r})')
        # Brokerage
        ws.cell(
            row=r,
            column=12,
            value=(
                f'=IF(G{r}="","",Settings!$C$5)'
            ),
        )
        # STT — only on SELL
        ws.cell(
            row=r,
            column=13,
            value=(
                f'=IF(G{r}="","",IF(F{r}="SELL",K{r}*IF(E{r}="Equity Intraday",'
                f'Settings!$C$8,Settings!$C$7),0))'
            ),
        )
        # Exchange + SEBI both sides
        ws.cell(
            row=r,
            column=14,
            value=f'=IF(G{r}="","",K{r}*(Settings!$C$9+Settings!$C$10))',
        )
        # Stamp on BUY
        ws.cell(
            row=r,
            column=15,
            value=f'=IF(G{r}="","",IF(F{r}="BUY",K{r}*Settings!$C$11,0))',
        )
        # GST on brokerage
        ws.cell(row=r, column=16, value=f'=IF(G{r}="","",L{r}*Settings!$C$6)')
        # DP on delivery SELL
        ws.cell(
            row=r,
            column=17,
            value=(
                f'=IF(G{r}="","",IF(AND(F{r}="SELL",E{r}="Equity Delivery"),Settings!$C$13,0))'
            ),
        )
        # Total charges
        ws.cell(row=r, column=18, value=f'=IF(G{r}="","",SUM(L{r}:Q{r}))')
        # Net cash flow: BUY negative, SELL positive
        ws.cell(
            row=r,
            column=19,
            value=(
                f'=IF(G{r}="","",IF(F{r}="BUY",-(K{r}+R{r}),K{r}-R{r}))'
            ),
        )
        for c in (7, 8, 11, 12, 13, 14, 15, 16, 17, 18, 19):
            ws.cell(row=r, column=c).number_format = MONEY_FMT
        ws.cell(row=r, column=2).number_format = "YYYY-MM-DD"

    # Data validation
    dv_ex = DataValidation(type="list", formula1='"NSE,BSE"', allow_blank=True)
    dv_seg = DataValidation(
        type="list",
        formula1='"Equity Delivery,Equity Intraday"',
        allow_blank=True,
    )
    dv_side = DataValidation(type="list", formula1='"BUY,SELL"', allow_blank=True)
    for dv in (dv_ex, dv_seg, dv_side):
        ws.add_data_validation(dv)
        dv.add(f"D{start_row}:D{max_row}")
    dv_seg.add(f"E{start_row}:E{max_row}")
    dv_side.add(f"F{start_row}:F{max_row}")

    ws.freeze_panes = "A2"
    set_col_widths(
        ws,
        {
            1: 8,
            2: 12,
            3: 14,
            4: 10,
            5: 18,
            6: 8,
            7: 8,
            8: 12,
            9: 12,
            10: 28,
            11: 14,
            12: 12,
            13: 10,
            14: 12,
            15: 10,
            16: 10,
            17: 8,
            18: 14,
            19: 16,
        },
    )


def build_open_positions(ws) -> None:
    ws.title = "Open Positions"
    ws["A1"] = "Open positions (net qty > 0 from Trade Log)"
    ws["A1"].font = TITLE_FONT
    headers = [
        "Symbol",
        "Exchange",
        "Segment",
        "Qty held",
        "Avg cost (₹)",
        "Cost value (₹)",
        "LTP (₹) — update manually",
        "Market value (₹)",
        "Unrealized P&L (₹)",
        "Unrealized %",
    ]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=3, column=j, value=h)
    style_header_row(ws, 3, len(headers))

    # Rows 4-53 for up to 50 symbols (user copies unique symbols or we use helper)
    # Use formulas referencing Trade Log — works for symbols listed in column A
    ws["A4"] = "List symbols you hold (one per row); sample:"
    ws["A5"] = "JINDALSTEL"
    for r in range(5, 55):
        ws.cell(row=r, column=2, value=f'=IF(A{r}="","",INDEX(\'Trade Log\'!D:D,MATCH(A{r},\'Trade Log\'!C:C,0)))')
        ws.cell(row=r, column=3, value=f'=IF(A{r}="","",INDEX(\'Trade Log\'!E:E,MATCH(A{r},\'Trade Log\'!C:C,0)))')
        buy_qty = f"SUMIFS('Trade Log'!G:G,'Trade Log'!C:C,A{r},'Trade Log'!F:F,\"BUY\")"
        sell_qty = f"SUMIFS('Trade Log'!G:G,'Trade Log'!C:C,A{r},'Trade Log'!F:F,\"SELL\")"
        ws.cell(row=r, column=4, value=f'=IF(A{r}="","",{buy_qty}-{sell_qty})')
        ws.cell(
            row=r,
            column=5,
            value=(
                f'=IF(D{r}<=0,"",SUMIFS(\'Trade Log\'!K:K,\'Trade Log\'!C:C,A{r},'
                f'\'Trade Log\'!F:F,"BUY")/SUMIFS(\'Trade Log\'!G:G,\'Trade Log\'!C:C,A{r},'
                f'\'Trade Log\'!F:F,"BUY"))'
            ),
        )
        ws.cell(row=r, column=6, value=f'=IF(D{r}<=0,"",D{r}*E{r})')
        ws.cell(row=r, column=8, value=f'=IF(D{r}<=0,"",D{r}*G{r})')
        ws.cell(row=r, column=9, value=f'=IF(D{r}<=0,"",H{r}-F{r})')
        ws.cell(row=r, column=10, value=f'=IF(F{r}<=0,"",I{r}/F{r})')
        for c in (4, 5, 6, 7, 8, 9):
            ws.cell(row=r, column=c).number_format = MONEY_FMT
        ws.cell(row=r, column=10).number_format = PCT_FMT

    ws["A57"] = "Totals"
    ws["F57"] = "=SUM(F5:F54)"
    ws["H57"] = "=SUM(H5:H54)"
    ws["I57"] = "=SUM(I5:I54)"
    for ref in ("F57", "H57", "I57"):
        ws[ref].number_format = MONEY_FMT
        ws[ref].font = Font(bold=True)

    set_col_widths(ws, {1: 14, 2: 10, 3: 18, 4: 10, 5: 14, 6: 14, 7: 18, 8: 14, 9: 16, 10: 12})
    ws.freeze_panes = "A4"


def build_closed_trades(ws) -> None:
    ws.title = "Closed Trades"
    ws["A1"] = "Round-trip trades (enter Position ID from Trade Log)"
    ws["A1"].font = TITLE_FONT
    headers = [
        "Position ID",
        "Symbol",
        "Segment",
        "Buy date",
        "Sell date",
        "Qty",
        "Buy avg (₹)",
        "Sell avg (₹)",
        "Gross P&L (₹)",
        "Total charges (₹)",
        "Net P&L (₹)",
        "Return %",
        "Hold days",
    ]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=3, column=j, value=h)
    style_header_row(ws, 3, len(headers))

    for r in range(4, 104):
        pid = f"A{r}"
        ws.cell(row=r, column=2, value=f'=IF({pid}="","",INDEX(\'Trade Log\'!C:C,MATCH({pid},\'Trade Log\'!I:I,0)))')
        ws.cell(row=r, column=3, value=f'=IF({pid}="","",INDEX(\'Trade Log\'!E:E,MATCH({pid},\'Trade Log\'!I:I,0)))')
        ws.cell(
            row=r,
            column=4,
            value=f'=IF({pid}="","",MINIFS(\'Trade Log\'!B:B,\'Trade Log\'!I:I,{pid},\'Trade Log\'!F:F,"BUY"))',
        )
        ws.cell(
            row=r,
            column=5,
            value=f'=IF({pid}="","",MAXIFS(\'Trade Log\'!B:B,\'Trade Log\'!I:I,{pid},\'Trade Log\'!F:F,"SELL"))',
        )
        ws.cell(
            row=r,
            column=6,
            value=(
                f'=IF({pid}="","",MIN(SUMIFS(\'Trade Log\'!G:G,\'Trade Log\'!I:I,{pid},'
                f'\'Trade Log\'!F:F,"BUY"),SUMIFS(\'Trade Log\'!G:G,\'Trade Log\'!I:I,{pid},'
                f'\'Trade Log\'!F:F,"SELL")))'
            ),
        )
        ws.cell(
            row=r,
            column=7,
            value=(
                f'=IF({pid}="","",SUMIFS(\'Trade Log\'!K:K,\'Trade Log\'!I:I,{pid},'
                f'\'Trade Log\'!F:F,"BUY")/SUMIFS(\'Trade Log\'!G:G,\'Trade Log\'!I:I,{pid},'
                f'\'Trade Log\'!F:F,"BUY"))'
            ),
        )
        ws.cell(
            row=r,
            column=8,
            value=(
                f'=IF({pid}="","",SUMIFS(\'Trade Log\'!K:K,\'Trade Log\'!I:I,{pid},'
                f'\'Trade Log\'!F:F,"SELL")/SUMIFS(\'Trade Log\'!G:G,\'Trade Log\'!I:I,{pid},'
                f'\'Trade Log\'!F:F,"SELL"))'
            ),
        )
        ws.cell(row=r, column=9, value=f'=IF({pid}="","",(H{r}-G{r})*F{r})')
        ws.cell(row=r, column=10, value=f'=IF({pid}="","",SUMIFS(\'Trade Log\'!R:R,\'Trade Log\'!I:I,{pid}))')
        ws.cell(row=r, column=11, value=f'=IF({pid}="","",I{r}-J{r})')
        ws.cell(row=r, column=12, value=f'=IF({pid}="","",IF(G{r}*F{r}=0,"",K{r}/(G{r}*F{r})))')
        ws.cell(row=r, column=13, value=f'=IF({pid}="","",E{r}-D{r})')
        for c in range(6, 12):
            ws.cell(row=r, column=c).number_format = MONEY_FMT
        ws.cell(row=r, column=12).number_format = PCT_FMT

    ws["A4"] = "T001"
    ws["A5"] = "T002"
    set_col_widths(ws, {1: 12, 2: 14, 3: 18, 4: 12, 5: 12, 6: 8, 7: 12, 8: 12, 9: 12, 10: 14, 11: 12, 12: 10, 13: 10})
    ws.freeze_panes = "A4"


def build_summary(ws) -> None:
    ws.title = "Summary"
    ws["A1"] = "Paper trading dashboard"
    ws["A1"].font = TITLE_FONT

    metrics = [
        ("Starting capital (₹)", "=Settings!C4"),
        ("Cash after all trades (₹)", "=Settings!C4+SUM('Trade Log'!S:S)"),
        ("Invested in open positions (₹)", "='Open Positions'!F57"),
        ("Portfolio value (cash + holdings)", "=B5+B6+'Open Positions'!H57"),
        ("Total realized P&L (₹)", "=SUM('Closed Trades'!K:K)"),
        ("Total unrealized P&L (₹)", "='Open Positions'!I57"),
        ("Total P&L (realized + unrealized)", "=B9+B10"),
        ("Return on starting capital %", "=IF(Settings!C4=0,\"\",B11/Settings!C4)"),
        ("Closed trades (count)", '=COUNTIF(\'Closed Trades\'!A4:A103,"?*")'),
        ("Winning trades", '=COUNTIF(\'Closed Trades\'!K4:K103,">0")'),
        ("Losing trades", '=COUNTIF(\'Closed Trades\'!K4:K103,"<0")'),
        ("Win rate %", "=IF(B14=0,\"\",B15/B14)"),
        ("Total brokerage paid (₹)", "=SUM('Trade Log'!L:L)"),
        ("Total charges paid (₹)", "=SUM('Trade Log'!R:R)"),
    ]
    row = 4
    ws.cell(row=row, column=1, value="Metric")
    ws.cell(row=row, column=2, value="Value")
    style_header_row(ws, row, 2)
    row += 1
    for label, formula in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=formula)
        if "₹" in label or "P&L" in label or "capital" in label or "Cash" in label or "Invested" in label or "Portfolio" in label or "brokerage" in label or "charges" in label:
            ws.cell(row=row, column=2).number_format = MONEY_FMT
        elif "%" in label or "rate" in label:
            ws.cell(row=row, column=2).number_format = PCT_FMT
        row += 1

    ws["A20"] = "Last updated: enter trades in Trade Log; update LTP on Open Positions for live MTM."
    ws["A20"].font = Font(italic=True, color="64748B")
    set_col_widths(ws, {1: 36, 2: 22})


def main() -> None:
    wb = Workbook()
    # Remove default sheet after creating ordered sheets
    default = wb.active
    build_instructions(wb.create_sheet("Instructions", 0))
    build_settings(wb.create_sheet("Settings", 1))
    build_trade_log(wb.create_sheet("Trade Log", 2))
    build_open_positions(wb.create_sheet("Open Positions", 3))
    build_closed_trades(wb.create_sheet("Closed Trades", 4))
    build_summary(wb.create_sheet("Summary", 5))
    wb.remove(default)

    wb.save(OUTPUT)
    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    main()
