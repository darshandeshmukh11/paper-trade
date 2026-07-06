"""Build and refresh paper_trading_india.xlsx (Excel-first paper trading)."""

from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from charges import ChargeSettings, compute_charges, net_cash_flow
from portfolio import completed_round_trips, trades_to_df

OUTPUT_NAME = "paper_trading_india.xlsx"
TRADE_LOG_MAX_ROW = 502
TRADE_DATA_START = 2

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1E3A5F")
MONEY_FMT = "#,##0.00"
PCT_FMT = "0.00%"
DATE_FMT = "YYYY-MM-DD"
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Trade Log columns
COL_DATE = 2
COL_SYMBOL = 3
COL_EXCHANGE = 4
COL_SEGMENT = 5
COL_SIDE = 6
COL_QTY = 7
COL_PRICE = 8
COL_POSITION_ID = 9
COL_NOTES = 10
COL_STOP = 11
COL_TARGET = 12
COL_GROSS = 13
COL_TOTAL_CHARGES = 20
COL_NET = 21
COL_CASH_AFTER = 22


def default_output_path() -> Path:
    return Path(__file__).resolve().parent / OUTPUT_NAME


def style_header_row(ws: Worksheet, row: int, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def set_col_widths(ws: Worksheet, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _charge_formulas(r: int) -> dict[int, str]:
    return {
        13: f'=IF(G{r}="","",G{r}*H{r})',
        14: f'=IF(G{r}="","",Settings!$C$5)',
        15: (
            f'=IF(G{r}="","",IF(F{r}="SELL",M{r}*IF(E{r}="Equity Intraday",'
            f"Settings!$C$8,Settings!$C$7),0))"
        ),
        16: f'=IF(G{r}="","",M{r}*(Settings!$C$9+Settings!$C$10))',
        17: f'=IF(G{r}="","",IF(F{r}="BUY",M{r}*Settings!$C$11,0))',
        18: f'=IF(G{r}="","",N{r}*Settings!$C$6)',
        19: f'=IF(G{r}="","",IF(AND(F{r}="SELL",E{r}="Equity Delivery"),Settings!$C$13,0))',
        20: f'=IF(G{r}="","",SUM(N{r}:S{r}))',
        21: f'=IF(G{r}="","",IF(F{r}="BUY",-(M{r}+T{r}),M{r}-T{r}))',
        22: f'=IF(G{r}="","",Settings!$C$4+SUM($U$2:U{r}))',
    }


def build_instructions(ws: Worksheet, output: Path) -> None:
    ws.title = "Instructions"
    lines = [
        "Paper trading — Excel workbook (NSE / BSE)",
        "",
        "Your file IS the database. Save to OneDrive / Google Drive / your Mac — no Streamlit backup needed.",
        "",
        "Daily workflow",
        "1. Settings — starting capital (₹) and broker charge assumptions.",
        "2. Trade Log — one row per fill (BUY or SELL). Leave Position ID blank for FIFO matching by symbol.",
        "3. Open Positions — auto lists symbols with net qty > 0 (Excel 365). Enter LTP in column G for mark-to-market.",
        "4. Completed Trades — run: python build_paper_trading_sheet.py --refresh",
        "   (after adding trades) to update realized P&L using the same FIFO logic as the web app.",
        "5. Dashboard — cash, equity, and performance at a glance.",
        "",
        "Import old JSON backup once:",
        "  python build_paper_trading_sheet.py --import ~/Downloads/paper_trading_backup.json",
        "",
        "Regenerate blank workbook:",
        "  python build_paper_trading_sheet.py",
        "",
        f"File: {output}",
    ]
    for i, line in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=line)
    ws["A1"].font = TITLE_FONT
    ws.column_dimensions["A"].width = 92


def build_settings(ws: Worksheet, starting_capital: float = 1_000_000) -> None:
    ws.title = "Settings"
    ws["A1"] = "Paper trading settings"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    rows = [
        ("", "Parameter", "Value", "Notes"),
        ("", "Starting capital (₹)", starting_capital, "Virtual cash"),
        ("", "Brokerage per order (₹)", 20, "Flat per order"),
        ("", "GST on brokerage (%)", 0.18, "18% on brokerage"),
        ("", "STT — delivery sell (%)", 0.001, "0.1% on sell turnover"),
        ("", "STT — intraday sell (%)", 0.00025, "0.025% on sell turnover"),
        ("", "Exchange txn charge (%)", 0.0000345, "NSE equity approx"),
        ("", "SEBI fee (%)", 0.000001, "₹10 per crore approx"),
        ("", "Stamp duty — buy (%)", 0.00015, "~0.015%"),
        ("", "DP charges — delivery sell (₹)", 15.93, "Per delivery sell"),
    ]
    for i, row in enumerate(rows, start=3):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    style_header_row(ws, 3, 4)
    set_col_widths(ws, {1: 4, 2: 32, 3: 18, 4: 42})
    ws["C4"].number_format = MONEY_FMT
    ws["C5"].number_format = MONEY_FMT
    for r in (7, 8, 9, 10, 11, 12):
        ws.cell(row=r, column=3).number_format = PCT_FMT
    ws["C13"].number_format = MONEY_FMT
    ws["A16"] = "Charge columns on Trade Log read Settings!$C$4:$C$13."
    ws["A16"].font = Font(italic=True, color="64748B")


def build_trade_log(ws: Worksheet, sample_rows: Optional[list[list[Any]]] = None) -> None:
    ws.title = "Trade Log"
    headers = [
        "#",
        "Date",
        "Symbol",
        "Exchange",
        "Segment",
        "Side",
        "Qty",
        "Price (₹)",
        "Position ID",
        "Notes",
        "Stop loss (₹)",
        "Target (₹)",
        "Gross (₹)",
        "Brokerage (₹)",
        "STT (₹)",
        "Exch+SEBI (₹)",
        "Stamp (₹)",
        "GST (₹)",
        "DP (₹)",
        "Total charges (₹)",
        "Net cash flow (₹)",
        "Cash after trade (₹)",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    samples = sample_rows or []
    for i, sample in enumerate(samples):
        r = TRADE_DATA_START + i
        for j, val in enumerate(sample, start=1):
            if j >= COL_GROSS:
                continue
            ws.cell(row=r, column=j, value=val)

    for r in range(TRADE_DATA_START, TRADE_LOG_MAX_ROW + 1):
        ws.cell(row=r, column=1, value=f'=IF(G{r}="","",ROW()-1)')
        for col, formula in _charge_formulas(r).items():
            ws.cell(row=r, column=col, value=formula)
        for c in range(COL_QTY, COL_NET + 1):
            ws.cell(row=r, column=c).number_format = MONEY_FMT
        ws.cell(row=r, column=COL_DATE).number_format = DATE_FMT

    dv_ex = DataValidation(type="list", formula1='"NSE,BSE"', allow_blank=True)
    dv_seg = DataValidation(type="list", formula1='"Equity Delivery,Equity Intraday"', allow_blank=True)
    dv_side = DataValidation(type="list", formula1='"BUY,SELL"', allow_blank=True)
    for dv in (dv_ex, dv_seg, dv_side):
        ws.add_data_validation(dv)
    dv_ex.add(f"D{TRADE_DATA_START}:D{TRADE_LOG_MAX_ROW}")
    dv_seg.add(f"E{TRADE_DATA_START}:E{TRADE_LOG_MAX_ROW}")
    dv_side.add(f"F{TRADE_DATA_START}:F{TRADE_LOG_MAX_ROW}")

    ws.freeze_panes = "A2"
    set_col_widths(
        ws,
        {
            1: 5,
            2: 12,
            3: 14,
            4: 10,
            5: 18,
            6: 8,
            7: 8,
            8: 12,
            9: 12,
            10: 24,
            11: 12,
            12: 12,
            13: 12,
            14: 10,
            15: 10,
            16: 12,
            17: 10,
            18: 10,
            19: 8,
            20: 14,
            21: 14,
            22: 16,
        },
    )


def build_open_positions(ws: Worksheet) -> None:
    ws.title = "Open Positions"
    ws["A1"] = "Open positions (net qty > 0)"
    ws["A1"].font = TITLE_FONT
    headers = [
        "Symbol",
        "Exchange",
        "Segment",
        "Qty held",
        "Avg cost (₹)",
        "Cost value (₹)",
        "LTP (₹) — enter manually",
        "Market value (₹)",
        "Unrealized P&L (₹)",
        "Unrealized %",
    ]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=3, column=j, value=h)
    style_header_row(ws, 3, len(headers))

    # Excel 365 dynamic array — spills down from A4
    spill_row = 4
    s = spill_row
    ws.cell(
        row=s,
        column=1,
        value=(
            f'=LET('
            f"log_syms,UNIQUE(FILTER('Trade Log'!$C$2:$C${TRADE_LOG_MAX_ROW},'Trade Log'!$G$2:$G${TRADE_LOG_MAX_ROW}<>\"\")),"
            f"nets,BYROW(log_syms,LAMBDA(sym,"
            f"SUMIFS('Trade Log'!$G$2:$G${TRADE_LOG_MAX_ROW},'Trade Log'!$C$2:$C${TRADE_LOG_MAX_ROW},sym,'Trade Log'!$F$2:$F${TRADE_LOG_MAX_ROW},\"BUY\")"
            f"-SUMIFS('Trade Log'!$G$2:$G${TRADE_LOG_MAX_ROW},'Trade Log'!$C$2:$C${TRADE_LOG_MAX_ROW},sym,'Trade Log'!$F$2:$F${TRADE_LOG_MAX_ROW},\"SELL\"))),"
            f"FILTER(log_syms,nets>0))"
        ),
    )
    r = s
    ws.cell(
        row=r,
        column=2,
        value=f'=IF(A{r}=\"\",\"\",INDEX(\'Trade Log\'!$D$2:$D${TRADE_LOG_MAX_ROW},MATCH(A{r},\'Trade Log\'!$C$2:$C${TRADE_LOG_MAX_ROW},0)))',
    )
    ws.cell(
        row=r,
        column=3,
        value=f'=IF(A{r}=\"\",\"\",INDEX(\'Trade Log\'!$E$2:$E${TRADE_LOG_MAX_ROW},MATCH(A{r},\'Trade Log\'!$C$2:$C${TRADE_LOG_MAX_ROW},0)))',
    )
    buy = (
        f"SUMIFS('Trade Log'!$G$2:$G${TRADE_LOG_MAX_ROW},'Trade Log'!$C$2:$C${TRADE_LOG_MAX_ROW},A{r},"
        f"'Trade Log'!$F$2:$F${TRADE_LOG_MAX_ROW},\"BUY\")"
    )
    sell = (
        f"SUMIFS('Trade Log'!$G$2:$G${TRADE_LOG_MAX_ROW},'Trade Log'!$C$2:$C${TRADE_LOG_MAX_ROW},A{r},"
        f"'Trade Log'!$F$2:$F${TRADE_LOG_MAX_ROW},\"SELL\")"
    )
    ws.cell(row=r, column=4, value=f"=IF(A{r}=\"\",\"\",{buy}-{sell})")
    ws.cell(
        row=r,
        column=5,
        value=(
            f"=IF(D{r}<=0,\"\",SUMIFS('Trade Log'!$M$2:$M${TRADE_LOG_MAX_ROW},'Trade Log'!$C$2:$C${TRADE_LOG_MAX_ROW},A{r},"
            f"'Trade Log'!$F$2:$F${TRADE_LOG_MAX_ROW},\"BUY\")/"
            f"SUMIFS('Trade Log'!$G$2:$G${TRADE_LOG_MAX_ROW},'Trade Log'!$C$2:$C${TRADE_LOG_MAX_ROW},A{r},"
            f"'Trade Log'!$F$2:$F${TRADE_LOG_MAX_ROW},\"BUY\"))"
        ),
    )
    ws.cell(row=r, column=6, value=f"=IF(D{r}<=0,\"\",D{r}*E{r})")
    ws.cell(row=r, column=8, value=f"=IF(D{r}<=0,\"\",D{r}*G{r})")
    ws.cell(row=r, column=9, value=f"=IF(D{r}<=0,\"\",H{r}-F{r})")
    ws.cell(row=r, column=10, value=f"=IF(F{r}<=0,\"\",I{r}/F{r})")
    for c in (4, 5, 6, 7, 8, 9):
        ws.cell(row=r, column=c).number_format = MONEY_FMT
    ws.cell(row=r, column=10).number_format = PCT_FMT

    ws["A56"] = "Totals"
    ws["F56"] = "=SUM(F4#)"
    ws["H56"] = "=SUM(H4#)"
    ws["I56"] = "=SUM(I4#)"
    for ref in ("F56", "H56", "I56"):
        ws[ref].number_format = MONEY_FMT
        ws[ref].font = Font(bold=True)

    ws["A58"] = (
        "Requires Excel 365 / Microsoft 365 for auto symbol list. "
        "Enter LTP in column G. Then run --refresh for exact realized P&L."
    )
    ws["A58"].font = Font(italic=True, color="64748B")
    set_col_widths(ws, {1: 14, 2: 10, 3: 18, 4: 10, 5: 14, 6: 14, 7: 20, 8: 14, 9: 16, 10: 12})
    ws.freeze_panes = "A4"


def build_dashboard(ws: Worksheet) -> None:
    ws.title = "Dashboard"
    ws["A1"] = "Paper trading dashboard"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Metric"
    ws["B3"] = "Value"
    style_header_row(ws, 3, 2)

    metrics = [
        ("Starting capital (₹)", "=Settings!C4"),
        ("Cash available (₹)", f"=Settings!C4+SUM('Trade Log'!$U${TRADE_DATA_START}:$U${TRADE_LOG_MAX_ROW})"),
        ("Invested in open positions (₹)", "='Open Positions'!F56"),
        ("Holdings at LTP (₹)", "='Open Positions'!H56"),
        ("Portfolio equity (₹)", "=B5+B8"),
        ("Unrealized P&L (₹)", "='Open Positions'!I56"),
        ("Realized P&L (₹)", "=SUM('Completed Trades'!H:H)"),
        ("Total P&L (₹)", "=B10+B9"),
        ("Return on starting capital %", "=IF(Settings!C4=0,\"\",B11/Settings!C4)"),
        ("Open positions (count)", "=COUNTA('Open Positions'!A4#)"),
        ("Completed round-trips", "=COUNTA('Completed Trades'!A2:A500)-COUNTIF('Completed Trades'!A2:A500,\"\")"),
        ("Winning trades", "=COUNTIF('Completed Trades'!H2:H500,\">0\")"),
        ("Win rate %", "=IF(B14=0,\"\",B15/B14)"),
        ("Total charges paid (₹)", f"=SUM('Trade Log'!$T${TRADE_DATA_START}:$T${TRADE_LOG_MAX_ROW})"),
    ]
    row = 4
    for label, formula in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=formula)
        if "₹" in label or "P&L" in label or "capital" in label or "Cash" in label or "Invested" in label or "Holdings" in label or "Portfolio" in label or "charges" in label:
            ws.cell(row=row, column=2).number_format = MONEY_FMT
        elif "%" in label or "rate" in label:
            ws.cell(row=row, column=2).number_format = PCT_FMT
        row += 1

    ws["A20"] = "Tip: save this file in OneDrive — edits persist; no cloud app backup required."
    ws["A20"].font = Font(italic=True, color="64748B")
    set_col_widths(ws, {1: 34, 2: 22})


def build_completed_trades_sheet(ws: Worksheet, trades_df: pd.DataFrame) -> None:
    ws.title = "Completed Trades"
    headers = [
        "Symbol",
        "Position ID",
        "Qty",
        "Buy avg (₹)",
        "Sell avg (₹)",
        "Buy date",
        "Sell date",
        "Net P&L (₹)",
        "Return %",
        "Charges (₹)",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    completed = completed_round_trips(trades_df)
    for _, row in completed.iterrows():
        buy_d = row["buy_date"]
        sell_d = row["sell_date"]
        if hasattr(buy_d, "date"):
            buy_d = buy_d.date() if isinstance(buy_d, datetime) else buy_d
        if hasattr(sell_d, "date"):
            sell_d = sell_d.date() if isinstance(sell_d, datetime) else sell_d
        ws.append(
            [
                row["symbol"],
                row.get("position_id") or "",
                int(row["qty"]),
                float(row["buy_avg"]),
                float(row["sell_avg"]),
                buy_d,
                sell_d,
                float(row["pnl_inr"]),
                float(row["return_pct"]) / 100.0,
                float(row.get("charges", 0)),
            ]
        )

    for r in range(2, ws.max_row + 1):
        for c in (3, 4, 5, 8, 10):
            ws.cell(row=r, column=c).number_format = MONEY_FMT
        ws.cell(row=r, column=9).number_format = PCT_FMT
        ws.cell(row=r, column=6).number_format = DATE_FMT
        ws.cell(row=r, column=7).number_format = DATE_FMT

    ws.freeze_panes = "A2"
    set_col_widths(ws, {1: 14, 2: 12, 3: 8, 4: 12, 5: 12, 6: 12, 7: 12, 8: 12, 9: 10, 10: 12})
    if ws.max_row < 2:
        ws["A2"] = "(Run build_paper_trading_sheet.py --refresh after logging trades)"


def _charge_settings_from_sheet(wb: Workbook) -> ChargeSettings:
    if "Settings" not in wb.sheetnames:
        return ChargeSettings()
    ws = wb["Settings"]
    return ChargeSettings(
        dp_delivery_sell=float(ws["C13"].value or 15.93),
    )


def trades_df_from_workbook(wb: Workbook) -> tuple[pd.DataFrame, float]:
    if "Trade Log" not in wb.sheetnames:
        return pd.DataFrame(), 500_000.0
    ws = wb["Trade Log"]
    settings_ws = wb["Settings"] if "Settings" in wb.sheetnames else None
    starting = float(settings_ws["C4"].value) if settings_ws and settings_ws["C4"].value else 1_000_000.0
    cs = _charge_settings_from_sheet(wb)

    rows: list[dict[str, Any]] = []
    for r in range(TRADE_DATA_START, TRADE_LOG_MAX_ROW + 1):
        qty = ws.cell(row=r, column=COL_QTY).value
        if qty in (None, "", 0):
            continue
        side = ws.cell(row=r, column=COL_SIDE).value
        symbol = ws.cell(row=r, column=COL_SYMBOL).value
        if not side or not symbol:
            continue

        traded = ws.cell(row=r, column=COL_DATE).value
        if isinstance(traded, datetime):
            traded_iso = traded.replace(tzinfo=None).isoformat()
        elif isinstance(traded, date):
            traded_iso = datetime.combine(traded, datetime.min.time()).isoformat()
        else:
            traded_iso = str(traded)

        segment = str(ws.cell(row=r, column=COL_SEGMENT).value or "Equity Delivery")
        exchange = str(ws.cell(row=r, column=COL_EXCHANGE).value or "NSE")
        price = float(ws.cell(row=r, column=COL_PRICE).value or 0)
        qty_i = int(qty)
        side_u = str(side).upper()

        gross_val = ws.cell(row=r, column=COL_GROSS).value
        charges_val = ws.cell(row=r, column=COL_TOTAL_CHARGES).value
        net_val = ws.cell(row=r, column=COL_NET).value

        if isinstance(gross_val, (int, float)) and isinstance(charges_val, (int, float)):
            gross = float(gross_val)
            charges = float(charges_val)
        else:
            ch = compute_charges(side_u, qty_i, price, segment, cs, exchange=exchange)
            gross = ch.gross
            charges = ch.total

        if isinstance(net_val, (int, float)):
            net = float(net_val)
        else:
            net = net_cash_flow(
                side_u,
                compute_charges(side_u, qty_i, price, segment, cs, exchange=exchange),
            )

        rows.append(
            {
                "id": r,
                "traded_at": traded_iso,
                "symbol": str(symbol).upper(),
                "exchange": ws.cell(row=r, column=COL_EXCHANGE).value or "NSE",
                "segment": segment,
                "side": side_u,
                "qty": qty_i,
                "price": price,
                "position_id": ws.cell(row=r, column=COL_POSITION_ID).value or None,
                "notes": ws.cell(row=r, column=COL_NOTES).value or "",
                "gross": gross,
                "charges": charges,
                "net_cash": net,
            }
        )

    return trades_to_df(rows), starting


def json_to_trade_rows(data: dict[str, Any]) -> list[list[Any]]:
    out: list[list[Any]] = []
    for t in sorted(data.get("trades", []), key=lambda x: (x.get("traded_at", ""), x.get("id", 0))):
        traded = t.get("traded_at", "")[:10]
        try:
            d = datetime.fromisoformat(traded.replace("Z", "+00:00")).date()
        except ValueError:
            d = traded
        out.append(
            [
                None,
                d,
                str(t.get("symbol", "")).upper(),
                t.get("exchange", "NSE"),
                t.get("segment", "Equity Delivery"),
                t.get("side", "BUY"),
                int(t.get("qty", 0)),
                float(t.get("price", 0)),
                t.get("position_id") or "",
                t.get("notes") or "",
                t.get("stop_loss") or "",
                t.get("target_price") or "",
            ]
        )
    return out


def starting_capital_from_json(data: dict[str, Any]) -> float:
    raw = data.get("settings", {}).get("starting_capital", "1000000")
    return float(raw)


def write_trade_rows_to_sheet(ws: Worksheet, rows: list[list[Any]]) -> None:
    for i, row in enumerate(rows):
        r = TRADE_DATA_START + i
        for j, val in enumerate(row, start=1):
            if j >= COL_GROSS:
                continue
            ws.cell(row=r, column=j, value=val)


def create_workbook(
    output: Path,
    *,
    import_json: Optional[dict[str, Any]] = None,
    sample_rows: Optional[list[list[Any]]] = None,
) -> None:
    wb = Workbook()
    default = wb.active

    starting = 1_000_000.0
    trade_rows = sample_rows
    if import_json:
        starting = starting_capital_from_json(import_json)
        trade_rows = json_to_trade_rows(import_json)

    build_instructions(wb.create_sheet("Instructions", 0), output)
    build_settings(wb.create_sheet("Settings", 1), starting_capital=starting)
    build_trade_log(wb.create_sheet("Trade Log", 2), sample_rows=trade_rows)
    build_open_positions(wb.create_sheet("Open Positions", 3))

    trades_df = pd.DataFrame()
    if import_json:
        trade_dicts = []
        for i, t in enumerate(
            sorted(
                import_json.get("trades", []),
                key=lambda x: (x.get("traded_at", ""), x.get("id", 0)),
            ),
            start=1,
        ):
            trade_dicts.append(
                {
                    **{k: v for k, v in t.items() if k != "id"},
                    "id": t.get("id", i),
                    "symbol": str(t.get("symbol", "")).upper(),
                    "side": str(t.get("side", "")).upper(),
                    "net_cash": float(t.get("net_cash", 0)),
                    "charges": float(t.get("charges", 0)),
                    "gross": float(t.get("gross", 0)),
                }
            )
        trades_df = trades_to_df(trade_dicts)

    build_completed_trades_sheet(wb.create_sheet("Completed Trades", 4), trades_df)
    build_dashboard(wb.create_sheet("Dashboard", 5))
    wb.remove(default)
    wb.save(output)


def refresh_workbook(path: Path) -> None:
    wb_values = load_workbook(path, data_only=True)
    trades_df, _ = trades_df_from_workbook(wb_values)
    wb = load_workbook(path)
    if "Completed Trades" in wb.sheetnames:
        del wb["Completed Trades"]
    build_completed_trades_sheet(wb.create_sheet("Completed Trades"), trades_df)
    wb.save(path)


def load_json_backup(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))
